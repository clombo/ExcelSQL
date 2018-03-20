"""
    TODO
    Make full cmd script with commands etc
    Directory where script is will be seen as the default dir
    File name needs to be specified

    Specify own directory

    Make robust to accept any format of spreadsheets
"""

import os
import argparse

#This module does not come with the python installation and needs to be
#installed
import openpyxl

#Error messages
class NoActionFound(Exception):
    pass

class WorkBookNotFound(Exception):
    pass

#SQL build function
def SQLBuildFromSheet(sheet,tb):

    tempval = ''
    finalval = '--{}\n'.format(tb)

    #row and col start
    r = 2
    c = 1

    #Build insert
    for c in range(1,sheet.max_column):
        
        #Get values and add to string
        tempval += str(sheet.cell(row=1,column=c).value)
        if c != sheet.max_column-1:
            tempval += ','

    #Add to final value and clear out temp string
    finalval += 'INSERT INTO '+tb+'('+tempval+')\n'
    tempval = ''

    #Build VALUES statements
    for r in range(2,sheet.max_row+1):
        for c in range(1,sheet.max_column):
            
            #Get values and add to string
            val = str(sheet.cell(row=r,column=c).value)
            if val == 'NULL':
                tempval += val
            else:
                tempval += "'"+val+"'"

            if c != sheet.max_column-1:
                tempval += ','

        finalval += 'VALUES('+tempval+')\n'
        tempval = ''

        return finalval

#Save file function
def SaveFile(filename,action,dir,content):

    """
    action
    1 -> New file
    2 -> append to file
    """
    if action == 1:
        #Save to new file
        resultFile = open(os.path.join(dir,filename),'w')
        resultFile.write(content)
        resultFile.close()
    elif action == 2:
        #Append to file if exists, if not then create it
        pass
    else:
        #Raise error
        raise NoActionFound('Action specified not found!')

def Main():
    #Create parser
    parser = argparse.ArgumentParser(description="Generate SQL code")


    parser.add_argument('table',help='Table name to use in SQL generation.',type=str)

    #Group arguments to either save to existing file (append) or save to new file (table name)
    group2 = parser.add_mutually_exclusive_group()
    group2.add_argument('--append','-a',help='append to default file pySQLGen', action='store_true')
    group2.add_argument('--new','-n',help='Add SQL to new file. File saved as table name',action='store_true')

    args = parser.parse_args()

    #Directory details
    defaultFile = 'pySQLGen.sql'
    WorkingDir = os.getcwd()

    #Workbook name
    wbName = '{}.xlsx'.format(args.table)

    if os.path.isfile(os.path.join(WorkingDir,wbName)):
        print 'Opening Workbook...'
        wb = openpyxl.load_workbook(os.path.join(WorkingDir,wbName), data_only=True)

        #Get sheet
        sheetname = wb.get_sheet_names()[0]
        sheet = wb.get_sheet_by_name(sheetname)

        #BuildSQL
        print 'Building SQL...'
        cont = SQLBuildFromSheet(sheet,args.table)

        print cont
    else:
        raise WorkBookNotFound('No Workbook Found!')


if __name__ == '__main__':
    print 'running...'
    Main()


"""
    

    #print finalval
    print 'Saving to file...'


    print 'DONE!'
    """
     
